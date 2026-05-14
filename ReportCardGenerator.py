import openpyxl
from collections import defaultdict
import shutil
from copy import copy
import os
from PySide6 import QtWidgets, QtGui, QtCore

class Worker(QtCore.QObject):
    finished = QtCore.Signal()
    error = QtCore.Signal(str)
    cleanup = QtCore.Signal()

    def __init__(self, func, args):
        super().__init__()
        self.func = func
        self.args = args

    def run(self):
        try:
            self.func(*self.args)
            self.finished.emit()
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.cleanup.emit()

# GUI
class MainWindow(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()

        self.setFixedSize(600, 200)
        self.setWindowTitle("Report Card Generator")

        def resource_path(filename):
            import sys, os
            if hasattr(sys, "_MEIPASS"):
                return os.path.join(sys._MEIPASS, filename)
            return os.path.join(os.path.abspath("."), filename)

        icon_path = resource_path("icon.png")
        self.setWindowIcon(QtGui.QIcon(icon_path))

        self.setWindowFlags(
            QtCore.Qt.Window |
            QtCore.Qt.WindowCloseButtonHint |
            QtCore.Qt.WindowMinimizeButtonHint
        )
        self.setWindowFlag(QtCore.Qt.WindowMaximizeButtonHint, False)

        self.grade_workbook_filepath = None
        self.report_card_template_filepath = None
        self.output_folder_filepath = None

        # Main vertical layout
        self.vlayout = QtWidgets.QVBoxLayout()

        # Loading bar
        self.progress = QtWidgets.QProgressBar()
        self.progress.setRange(0, 0)
        self.progress.hide()
        self.vlayout.addWidget(self.progress)

        # Horizontal layout 1 (grade)
        self.hlayout1 = QtWidgets.QHBoxLayout()

        self.button1 = QtWidgets.QPushButton("Select Grade Workbook")
        self.selected_grade_workbook_filepath = QtWidgets.QLabel(f"No file selected")

        self.button1.clicked.connect(self.select_grade_workbook)

        self.hlayout1.addWidget(self.button1)
        self.hlayout1.addWidget(self.selected_grade_workbook_filepath)

        # Horizontal layout 2 (report card template)
        self.hlayout2 = QtWidgets.QHBoxLayout()

        self.button2 = QtWidgets.QPushButton("Select Report Card Template")
        self.selected_report_card_template_filepath = QtWidgets.QLabel(f"No file selected")

        self.button2.clicked.connect(self.select_report_card_template)

        self.hlayout2.addWidget(self.button2)
        self.hlayout2.addWidget(self.selected_report_card_template_filepath)

        # Horizontal layout 3 (output folder selection)
        self.hlayout3 = QtWidgets.QHBoxLayout()

        self.button3 = QtWidgets.QPushButton("Select Output Folder")
        self.selected_output_folder_filepath = QtWidgets.QLabel(f"No folder selected")

        self.button3.clicked.connect(self.select_output_folder)

        self.hlayout3.addWidget(self.button3)
        self.hlayout3.addWidget(self.selected_output_folder_filepath)

        # Horizontal layout 4 (output folder name, filename)
        self.hlayout4 = QtWidgets.QHBoxLayout()
        
        self.textbox1 = QtWidgets.QLineEdit()
        self.textbox1.setPlaceholderText("Enter Output Folder Name (Optional)")
        self.textbox2 = QtWidgets.QLineEdit()
        self.textbox2.setPlaceholderText("Enter Filename Suffix (Optional)")

        self.hlayout4.addWidget(self.textbox1)
        self.hlayout4.addWidget(self.textbox2)

        # Horizontal layout 5 (generate button)
        self.hlayout5 = QtWidgets.QHBoxLayout()

        self.button4 = QtWidgets.QPushButton("Generate")
        self.button4.clicked.connect(self.generate_reports)

        self.hlayout5.addWidget(self.button4)

        # Add row to main layout
        self.vlayout.addLayout(self.hlayout1)
        self.vlayout.addLayout(self.hlayout2)
        self.vlayout.addLayout(self.hlayout3)
        self.vlayout.addLayout(self.hlayout4)
        self.vlayout.addLayout(self.hlayout5)

        self.setLayout(self.vlayout)

    def select_grade_workbook(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Grade Workbook")
        if path:
            self.selected_grade_workbook_filepath.setText(f"Selected: {path}")
            self.grade_workbook_filepath = path
    
    def select_report_card_template(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Report Card Template")
        if path:
            self.selected_report_card_template_filepath.setText(f"Selected: {path}")
            self.report_card_template_filepath = path

    def select_output_folder(self):
        path = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if path:
            self.selected_output_folder_filepath.setText(f"Selected: {path}")
            self.output_folder_filepath = path

    def generate_reports(self):
        if not self.grade_workbook_filepath or not self.report_card_template_filepath or not self.output_folder_filepath:
            QtWidgets.QMessageBox.warning(self, "Error", "Please select all required files/folders.")
            return
        self.thread = QtCore.QThread()
        self.worker = Worker(
            generate_report_cards,
            [
            self.grade_workbook_filepath,
            self.report_card_template_filepath,
            self.output_folder_filepath,
            self.textbox1.text(),
            self.textbox2.text()
            ]
        )
        self.worker.moveToThread(self.thread)

        self.worker.finished.connect(self.on_finished)
        self.worker.error.connect(self.on_error)

        self.worker.cleanup.connect(self.thread.quit)
        self.worker.cleanup.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)

        self.progress.show()
        self.button4.setEnabled(False)

        self.thread.start()
        self.thread.started.connect(self.worker.run)
    
    def on_error(self, message):
        self.reset_ui()
        QtWidgets.QMessageBox.critical(
            self,
            "Error",
            f"Something went wrong:\n\n{message}"
        )

    def reset_ui(self):
        self.progress.hide()
        self.button4.setEnabled(True)

    def on_finished(self):
        self.reset_ui()
        QtWidgets.QMessageBox.information(self, "Success", "Report cards generated successfully")

    def showFullScreen(self):
        pass

    def changeEvent(self, event):
        if self.windowState() & QtCore.Qt.WindowMaximized:
            self.setWindowState(QtCore.Qt.WindowNoState)
        super().changeEvent(event)

def generate_report_cards(
    grade_workbook_filepath,
    report_card_template_filepath,
    output_folder,
    folder_name,
    optional_text
):

    grade_wb = openpyxl.load_workbook(grade_workbook_filepath, data_only=True)

    # Input worksheet to get poi coords (function cannot be generalised; otherwise inefficient as each term would need to search the entire excel)
    def find_grade_worksheet_coordinates(worksheet):
        sid_coords = {}
        student_name_coords = {}
        batch_number_coords = {}
        subject_coords = {}
        total_marks_coords = {}
        grade_coords = {}
        for row in worksheet.iter_rows():
            for cell in row:
                value = str(cell.value).strip().lower() if cell.value else ""
                
                if value == "module":
                    subject_coords.update({"row": cell.row, "column": cell.column + 1})
                elif value == "sid":
                    sid_coords.update({"row": cell.row + 1, "column": cell.column})
                elif value == "student name":
                    student_name_coords.update({"row": cell.row + 1, "column": cell.column})
                elif value == "batch no.":
                    batch_number_coords.update({"row": cell.row + 1, "column": cell.column})
                elif value == "total (100 marks)":
                    total_marks_coords.update({"row": cell.row + 1, "column": cell.column})
                elif value == "grade":
                    grade_coords.update({"row": cell.row + 1, "column": cell.column})
                
            if sid_coords and student_name_coords and batch_number_coords and subject_coords and total_marks_coords and grade_coords:
                break
        return sid_coords, student_name_coords, batch_number_coords, subject_coords, total_marks_coords, grade_coords

    # Input cell coordinates to get cell value
    def get_grade_worksheet_data(worksheet, coords):
        cell_value = (worksheet.cell(
            row=coords["row"],
            column=coords["column"]
        ).value)
        return cell_value if cell_value is not None else "NA"

    # Initialise database dictionary
    database = defaultdict(
        lambda: {
            "student_name": "",
            "batch_number": "",
            "subject_list": {

                }
            }
        )

    # Loop to fill database up (each cycle == one subject)
    for worksheet in grade_wb.worksheets:
        # Coords of poi for one subject
        sid_coords, student_name_coords, batch_number_coords, subject_coords, total_marks_coords, grade_coords = find_grade_worksheet_coordinates(worksheet)
        subject_name = get_grade_worksheet_data(worksheet, subject_coords)

        for row in worksheet.iter_rows(min_row=sid_coords["row"]):
            sid = row[sid_coords["column"]-1].value

            if not sid:
                break

            # Write one student's detail into database
            database[sid]["student_name"] = get_grade_worksheet_data(worksheet, student_name_coords).title()
            database[sid]["batch_number"] = get_grade_worksheet_data(worksheet, batch_number_coords)
            database[sid]["subject_list"][subject_name.title()] = {
                "total_marks": get_grade_worksheet_data(worksheet, total_marks_coords),
                "grade": get_grade_worksheet_data(worksheet, grade_coords)
            }

            for counter in [sid_coords, student_name_coords, batch_number_coords, total_marks_coords, grade_coords]:
                counter["row"] += 1

    # Input worksheet to get poi coords (function cannot be generalised; otherwise inefficient as each term would need to search the entire excel)
    def find_template_worksheet_coordinates(worksheet):
        sid_coords = {}
        student_name_coords = {}
        batch_number_coords = {}
        subject_coords = {}
        total_marks_coords = {}
        grade_coords = {}
        for row in worksheet.iter_rows():
            for cell in row:
                value = str(cell.value).strip().lower() if cell.value else ""
                
                if value == "subject":
                    subject_coords.update({"row": cell.row + 3, "column": cell.column})
                elif value == "student id":
                    sid_coords.update({"row": cell.row, "column": cell.column + 1})
                elif value == "name":
                    student_name_coords.update({"row": cell.row, "column": cell.column + 1})
                elif value == "batch number":
                    batch_number_coords.update({"row": cell.row, "column": cell.column + 1})
                
            if sid_coords and student_name_coords and batch_number_coords and subject_coords:
                break
        total_marks_coords.update({"row": subject_coords["row"], "column": subject_coords["column"] + 3})
        grade_coords.update({"row": subject_coords["row"], "column": subject_coords["column"] + 4})
        return sid_coords, student_name_coords, batch_number_coords, subject_coords, total_marks_coords, grade_coords

    # Function to clone one row of subject, total marks and grade for duplication
    def clone_cell_style(source_cell, target_cell):
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    def insert_rows_full_shift(ws, row_idx, amount=1):
        merged_ranges = list(ws.merged_cells.ranges)
        ranges_to_move = []
        for m_range in merged_ranges:
            if m_range.min_row >= row_idx:
                ranges_to_move.append(m_range)
                ws.unmerge_cells(str(m_range))
        
        row_heights_to_move = []
        for r, dim in ws.row_dimensions.items():
            if r >= row_idx:
                row_heights_to_move.append((r, dim.height))
        
        for r, _ in row_heights_to_move:
            del ws.row_dimensions[r]

        ws.insert_rows(row_idx, amount)
        
        for old_row, height in row_heights_to_move:
            new_row = old_row + amount
            ws.row_dimensions[new_row].height = height

        for m_range in ranges_to_move:
            m_range.shift(row_shift=amount, col_shift=0)
            ws.merge_cells(str(m_range))
            
        for image in ws._images:
            if (image.anchor._from.row + 1) >= row_idx:
                image.anchor._from.row += amount
                if image.anchor.to:
                    image.anchor.to.row += amount

    if folder_name:
        output_folder = os.path.join(output_folder, folder_name)
    else:
        output_folder = os.path.join(output_folder, "Report Card Folder")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    template_wb = openpyxl.load_workbook(report_card_template_filepath, data_only=True)
    template_ws = template_wb.active  
    sid_coords, student_name_coords, batch_number_coords, subject_coords, total_marks_coords, grade_coords = find_template_worksheet_coordinates(template_ws)

    for sid, info in database.items():
        filename = f"{info['student_name']} {optional_text}".strip() + ".xlsx"
        file_path = os.path.join(output_folder, filename)
        shutil.copy(report_card_template_filepath, file_path)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        ws.cell(row=sid_coords["row"], column=sid_coords["column"]).value = sid
        ws.cell(row=student_name_coords["row"], column=student_name_coords["column"]).value = info["student_name"]
        ws.cell(row=batch_number_coords["row"], column=batch_number_coords["column"]).value = info["batch_number"]

        start_row = subject_coords["row"]
        for i, (subject_name, subject_info) in enumerate(info["subject_list"].items()):
            current_row = start_row + i
            if i > 0:
                insert_rows_full_shift(ws, current_row)
                ws.row_dimensions[current_row].height = ws.row_dimensions[start_row].height

                for column_counter in range(1,15):
                    source = ws.cell(row=start_row, column=column_counter)
                    target = ws.cell(row=current_row, column=column_counter)
                    clone_cell_style(source, target)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

            ws.cell(row=current_row, column=subject_coords["column"]).value = subject_name
            ws.cell(row=current_row, column=total_marks_coords["column"]).value = subject_info["total_marks"]
            ws.cell(row=current_row, column=grade_coords["column"]).value = subject_info["grade"]
        wb.save(file_path)

app = QtWidgets.QApplication([])
window = MainWindow()
window.show()
app.exec()